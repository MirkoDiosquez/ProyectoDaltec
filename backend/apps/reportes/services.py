from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.core.exceptions import PermissionDenied
from django.core.files.base import ContentFile
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from apps.acciones.models import EstadoAccion
from apps.hallazgos.models import Hallazgo
from apps.reportes.models import ReporteHallazgos


REPORT_HEADERS = [
    "ID Hallazgo",
    "Fecha Creacion",
    "Tipo",
    "Estado",
    "Sector",
    "Subseccion",
    "Descripcion",
    "Ubicacion",
    "Creado Por",
    "Cliente Asociado",
    "Responsables",
    "Acciones Totales",
    "Acciones Cerradas",
    "Acciones Pendientes",
]


class ReporteHallazgosService:
    @staticmethod
    def _map_tipo_excel(tipo):
        mapping = {
            "NO_CONFORMIDAD": "NC",
            "OPORTUNIDAD_MEJORA": "OM",
            "QUEJA_CLIENTE": "QC",
        }
        return mapping.get(tipo, tipo or "")

    @staticmethod
    def _map_sector_excel(hallazgo):
        if not hallazgo.sector:
            return ""
        codigo = (hallazgo.sector.codigo or "").upper()
        if codigo == "RECLAMO_CLIENTE":
            return "RECLAMO"
        return codigo or (hallazgo.sector.nombre or "")

    @staticmethod
    def _map_estado_excel(hallazgo, acciones):
        if hallazgo.estado == "CERRADO":
            return "CERRADA"
        if not acciones:
            return "SIN TRATAR"

        total = len(acciones)
        cerradas = sum(1 for a in acciones if a.estado == EstadoAccion.CERRADA)
        if cerradas == 0:
            return "SIN TRATAR"
        if cerradas == total:
            return "CERRADA"
        return "EN TRATAMIENTO"

    @staticmethod
    def _build_template_rows():
        hallazgos = (
            Hallazgo.objects.select_related(
                "creado_por",
                "cliente_asociado",
                "sector",
                "subseccion",
            )
            .prefetch_related("responsables", "acciones", "porques")
            .order_by("-fecha_creacion", "-id")
        )

        rows = []
        for hallazgo in hallazgos:
            acciones = list(hallazgo.acciones.all())

            accion_inmediata = next((a for a in acciones if a.tipo == "INMEDIATA"), None)
            accion_correctiva = next((a for a in acciones if a.tipo == "CORRECTIVA"), None)
            accion_verificacion = next((a for a in acciones if a.tipo == "VERIFICACION_EFICIENCIA"), None)

            porques = list(hallazgo.porques.all().order_by("-created_at")) if hasattr(hallazgo, "porques") else []
            analisis = "\n".join(p.texto_causa for p in porques[:5] if p.texto_causa)

            responsables = ", ".join(
                f"{u.nombre} {u.apellido}".strip()
                for u in hallazgo.responsables.all()
            )

            rows.append([
                ReporteHallazgosService._map_tipo_excel(hallazgo.tipo),  # A Tipo
                ReporteHallazgosService._map_sector_excel(hallazgo),  # B Tipo (sector)
                hallazgo.id,  # C Nro.
                hallazgo.fecha_creacion,  # D Fecha
                (
                    f"{hallazgo.cliente_asociado.nombre} {hallazgo.cliente_asociado.apellido}".strip()
                    if hallazgo.cliente_asociado
                    else (hallazgo.subseccion.nombre if hallazgo.subseccion else "")
                ),  # E Cliente/Sector/Proveedor/Proceso
                "",  # F Documento asociado
                hallazgo.descripcion or "",  # G Hallazgo
                analisis,  # H Analisis 5 porques
                accion_inmediata.descripcion if accion_inmediata else "",  # I Accion inmediata
                accion_inmediata.fecha_fin if accion_inmediata else None,  # J Cierre
                accion_correctiva.descripcion if accion_correctiva else "",  # K Accion correctiva
                accion_correctiva.fecha_fin if accion_correctiva else None,  # L Cierre
                accion_verificacion.descripcion if accion_verificacion else "",  # M Verificacion eficacia
                accion_verificacion.fecha_fin if accion_verificacion else None,  # N Cierre
                "",  # O Costos
                "",  # P Comentarios/Obs/Fotos
                ReporteHallazgosService._map_estado_excel(hallazgo, acciones),  # Q Estado
                responsables,  # R Responsable actual
                "",  # S Riesgo operacional
                "",  # T Reservado
            ])

        return rows

    @staticmethod
    def _update_indicadores_sheet(workbook):
        if "INDICADORES" not in workbook.sheetnames:
            return

        ws = workbook["INDICADORES"]

        # NC por tipo (gráfico chart1)
        ws["B2"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!B4:B1048576,"INTERNO")'
        ws["B3"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!B4:B1048576,"PROVEEDOR")'
        ws["B4"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!B4:B1048576,"RECLAMO")'

        # OM/QC por tipo (gráfico chart2, manteniendo labels de plantilla)
        ws["E2"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"QC")'
        ws["E3"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!B4:B1048576,"INTERNO")'
        ws["E4"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!B4:B1048576,"PROVEEDOR")'

        # Estado NC (gráfico chart3)
        ws["B6"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!Q4:Q1048576,"CERRADA")'
        ws["B7"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!Q4:Q1048576,"VERIFICACIÓN DE EFICACIA")'
        ws["B8"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!Q4:Q1048576,"EN TRATAMIENTO")'
        ws["B9"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"NC",\'NO CONFORMIDADES\'!Q4:Q1048576,"SIN TRATAR")'

        # Estado OM/QC (gráfico chart4)
        ws["E6"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!Q4:Q1048576,"CERRADA")+COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"QC",\'NO CONFORMIDADES\'!Q4:Q1048576,"CERRADA")'
        ws["E7"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!Q4:Q1048576,"VERIFICACIÓN DE EFICACIA")+COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"QC",\'NO CONFORMIDADES\'!Q4:Q1048576,"VERIFICACIÓN DE EFICACIA")'
        ws["E8"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!Q4:Q1048576,"EN TRATAMIENTO")+COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"QC",\'NO CONFORMIDADES\'!Q4:Q1048576,"EN TRATAMIENTO")'
        ws["E9"] = '=COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"OM",\'NO CONFORMIDADES\'!Q4:Q1048576,"SIN TRATAR")+COUNTIFS(\'NO CONFORMIDADES\'!A4:A1048576,"QC",\'NO CONFORMIDADES\'!Q4:Q1048576,"SIN TRATAR")'

    @staticmethod
    def _build_rows():
        hallazgos = (
            Hallazgo.objects.select_related(
                "creado_por",
                "cliente_asociado",
                "sector",
                "subseccion",
            )
            .prefetch_related("responsables", "acciones")
            .order_by("-fecha_creacion", "-id")
        )

        rows = []
        for hallazgo in hallazgos:
            acciones = list(hallazgo.acciones.all())
            total_acciones = len(acciones)
            cerradas = sum(1 for accion in acciones if accion.estado == EstadoAccion.CERRADA)
            pendientes = total_acciones - cerradas

            responsables = ", ".join(
                f"{u.nombre} {u.apellido}".strip()
                for u in hallazgo.responsables.all()
            )

            rows.append([
                hallazgo.id,
                hallazgo.fecha_creacion.strftime("%Y-%m-%d"),
                hallazgo.tipo,
                hallazgo.estado,
                hallazgo.sector.nombre if hallazgo.sector else "",
                hallazgo.subseccion.nombre if hallazgo.subseccion else "",
                hallazgo.descripcion,
                hallazgo.ubicacion,
                f"{hallazgo.creado_por.nombre} {hallazgo.creado_por.apellido}".strip(),
                (
                    f"{hallazgo.cliente_asociado.nombre} {hallazgo.cliente_asociado.apellido}".strip()
                    if hallazgo.cliente_asociado
                    else ""
                ),
                responsables,
                total_acciones,
                cerradas,
                pendientes,
            ])

        return rows

    @staticmethod
    def _get_template_path():
        configured = getattr(
            settings,
            "REPORTE_HALLAZGOS_TEMPLATE",
            "TC-SGC-03_NC-OM_rev.06C.xlsm",
        )
        path = Path(configured)
        if path.is_absolute():
            return path

        backend_candidate = Path(settings.BASE_DIR) / configured
        if backend_candidate.exists():
            return backend_candidate

        # Allow templates stored at repository root (one level above backend/).
        project_root_candidate = Path(settings.BASE_DIR).parent / configured
        return project_root_candidate

    @staticmethod
    def _render_workbook(rows):
        template_path = ReporteHallazgosService._get_template_path()
        start_row = int(getattr(settings, "REPORTE_HALLAZGOS_START_ROW", 2))

        if template_path.exists():
            workbook = load_workbook(template_path, keep_vba=True)
            worksheet = workbook["NO CONFORMIDADES"] if "NO CONFORMIDADES" in workbook.sheetnames else workbook.active

            data_start_row = 4
            max_data_row = max(worksheet.max_row, data_start_row)

            # Clear previous dataset values while preserving template formatting/styles.
            for row_idx in range(data_start_row, max_data_row + 1):
                for col_idx in range(1, 21):  # A:T
                    worksheet.cell(row=row_idx, column=col_idx, value=None)

            write_row = data_start_row
            template_rows = ReporteHallazgosService._build_template_rows()
            for row_values in template_rows:
                for col_index, value in enumerate(row_values, start=1):
                    worksheet.cell(row=write_row, column=col_index, value=value)
                write_row += 1

            ReporteHallazgosService._update_indicadores_sheet(workbook)

            extension = ".xlsm"
        else:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = "Reporte Hallazgos"
            worksheet.append(REPORT_HEADERS)
            for row_values in rows:
                worksheet.append(row_values)
            extension = ".xlsx"

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output, extension

    @staticmethod
    def generar_reporte(admin_user):
        if not getattr(admin_user, "is_admin", False):
            raise PermissionDenied("Solo administradores pueden generar reportes.")

        rows = ReporteHallazgosService._build_rows()
        workbook_content, extension = ReporteHallazgosService._render_workbook(rows)

        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"reporte_hallazgos_{timestamp}{extension}"

        reporte = ReporteHallazgos.objects.create(
            nombre=filename,
            creado_por=admin_user,
        )
        reporte.archivo.save(filename, ContentFile(workbook_content.getvalue()), save=True)
        return reporte

    @staticmethod
    def eliminar_reporte(reporte):
        if reporte.archivo:
            reporte.archivo.delete(save=False)
        reporte.delete()
